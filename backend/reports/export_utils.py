import html
import os
import uuid

import openpyxl
from openpyxl.styles import Font, PatternFill
from django.conf import settings

from common.pdf_render import html_to_pdf_file

def generate_excel_file(report_type, headers, data_rows):
    """
    Generates an Excel file and saves it to the media reports directory.
    Returns the relative URL of the saved file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(report_type).title().replace('_', ' ')[:31]

    # Styling for header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row_num, row_data in enumerate(data_rows, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Adjust auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Prepare file directory
    reports_dir = os.path.join(settings.BASE_DIR, 'media', 'exports')
    os.makedirs(reports_dir, exist_ok=True)
    
    file_name = f"{report_type}_{uuid.uuid4().hex[:8]}.xlsx"
    file_path = os.path.join(reports_dir, file_name)
    
    wb.save(file_path)
    
    # Return relative URL
    return f"/media/exports/{file_name}"


def generate_pdf_file(report_type, headers, data_rows):
    """
    Render tabular report data to PDF via shared WeasyPrint helper (same folder as Excel exports).
    """
    title = str(report_type).replace('_', ' ')
    esc = html.escape
    th = ''.join(f'<th>{esc(str(h))}</th>' for h in headers)
    body_rows = []
    for row in data_rows:
        cells = ''.join(f'<td>{esc(str(v))}</td>' for v in row)
        body_rows.append(f'<tr>{cells}</tr>')
    html_string = f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8">
<style>
body {{ font-family: DejaVu Sans, Helvetica, Arial, sans-serif; font-size: 10pt; margin: 24px; }}
h1 {{ font-size: 14pt; color: #1e1b4b; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 12px; }}
th, td {{ border: 1px solid #cbd5e1; padding: 6px 8px; text-align: left; vertical-align: top; }}
th {{ background: #4f46e5; color: #fff; }}
tr:nth-child(even) td {{ background: #f8fafc; }}
</style>
</head>
<body>
<h1>{esc(title)}</h1>
<table>
<thead><tr>{th}</tr></thead>
<tbody>{''.join(body_rows)}</tbody>
</table>
</body>
</html>"""
    reports_dir = os.path.join(settings.BASE_DIR, 'media', 'exports')
    os.makedirs(reports_dir, exist_ok=True)
    file_name = f"{report_type}_{uuid.uuid4().hex[:8]}.pdf"
    file_path = os.path.join(reports_dir, file_name)
    html_to_pdf_file(html_string, file_path)
    return f"/media/exports/{file_name}"
